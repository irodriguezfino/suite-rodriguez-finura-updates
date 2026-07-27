from __future__ import annotations

import tempfile
import unittest
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from suite_pyside6.core.reparto_merma_precintos import (
    AX_CSV_FORMAT,
    AXCsvFormat,
    DomainValidationError,
    build_preview,
    calculate_adjustment,
    read_source_file,
    read_fac_files,
    render_ax_csv,
    render_ax_csv_records,
    write_ax_csv,
    validate_ax_csv_content,
    validate_final_weight,
    validate_work_order,
)


class RepartoMermaPrecintosTests(unittest.TestCase):
    def read(self, messages: list[str], include_header: bool = True):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "origen.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Mensajes"
            if include_header:
                worksheet.append(["Mensaje (09:40:14)"])
            for message in messages:
                cell = worksheet.cell(worksheet.max_row + 1, 1, message)
                cell.data_type = "s"
            workbook.save(path)
            workbook.close()
            return read_source_file(path)

    def result(self, messages: list[str], final_weight: str):
        return calculate_adjustment(self.read(messages), final_weight)

    def test_lee_el_formato_real_de_mensajes_excel(self):
        source = self.read([
            "1552803101921;WB065;12,00;1,00",
            "1452803106943;WB065;13,30;1,00",
            "1452803109104;WB065;14,10;1,00",
            "1552803112156;WB065;15,10;1.00",
            "1452803101993;WB065;16,20;1,00",
            "1652701110886;F00568;17,50;1,00",
        ])
        self.assertTrue(source.is_valid)
        self.assertEqual(source.source_format.worksheet, "Mensajes")
        self.assertEqual(source.source_format.column, "A")
        self.assertTrue(source.source_format.has_message_header)
        self.assertEqual([record.precinto for record in source.records], [
            "1552803101921", "1452803106943", "1452803109104",
            "1552803112156", "1452803101993", "1652701110886",
        ])
        self.assertEqual([record.peso_original for record in source.records], [
            Decimal("12.00"), Decimal("13.30"), Decimal("14.10"),
            Decimal("15.10"), Decimal("16.20"), Decimal("17.50"),
        ])
        self.assertEqual(source.total_weight, Decimal("88.20"))

    def test_lee_formato_actual_ibericos_con_mensaje_en_columna_c(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "precintos_ibericos.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Hoja1"
            worksheet.append(["Información", "Mensaje (13:25:44)", "157587958924;IB4200FCA0000;12,34;1,00    29/01/2026"])
            worksheet.append(["Información", "Mensaje (13:25:44)", "157587959273;IB4200FCA0000;12,36;1,00    29/01/2026"])
            workbook.save(path)
            workbook.close()
            source = read_source_file(path)
        self.assertTrue(source.is_valid)
        self.assertEqual(source.source_format.worksheet, "Hoja1")
        self.assertEqual(source.source_format.column, "C")
        self.assertTrue(source.source_format.has_message_header)
        self.assertEqual([record.precinto for record in source.records], ["157587958924", "157587959273"])
        self.assertEqual(source.total_weight, Decimal("24.70"))

    def test_calculos_normales_ausencia_merma_y_ganancia(self):
        messages = ["A;X;10,00;1,00", "B;X;20,00;1,00"]
        for final, loss, percentage, factor in (
            ("0,00", Decimal("30.00"), Decimal("100"), Decimal("0")),
            ("27,00", Decimal("3.00"), Decimal("10.0"), Decimal("0.9")),
            ("30,00", Decimal("0.00"), Decimal("0"), Decimal("1")),
            ("33,00", Decimal("-3.00"), Decimal("-10.0"), Decimal("1.1")),
        ):
            with self.subTest(final=final):
                result = self.result(messages, final)
                self.assertEqual(result.absolute_loss, loss)
                self.assertEqual(result.loss_percentage, percentage)
                self.assertEqual(result.adjustment_factor, factor)
                self.assertEqual(result.adjusted_total, Decimal(final.replace(",", ".")))

    def test_un_precinto_varios_precintos_y_decimales(self):
        one = self.result(["UNO;X;10,25;1,00"], "9,99")
        self.assertEqual(one.rows[0].adjusted_weight, Decimal("9.99"))
        many = self.result(["A;X;1,11;1", "B;X;2,22;1", "C;X;6,67;1"], "8,50")
        self.assertEqual([row.source.precinto for row in many.rows], ["A", "B", "C"])
        self.assertEqual(many.adjusted_total, Decimal("8.50"))

    def test_coma_punto_sin_avisos(self):
        comma = self.read(["A;X;1,25;1"])
        point = self.read(["A;X;1.25;1"])
        mixed = self.read(["A;X;1,25;1", "B;X;2.50;1"])
        self.assertEqual(comma.total_weight, Decimal("1.25"))
        self.assertEqual(point.total_weight, Decimal("1.25"))
        self.assertTrue(mixed.is_valid)
        self.assertEqual(mixed.issues, ())

    def test_residuo_positivo_y_negativo_se_concilian(self):
        messages = ["A;X;1;1", "B;X;1;1", "C;X;1;1"]
        positive = self.result(messages, "1,00")
        negative = self.result(messages, "2,00")
        self.assertEqual([row.adjusted_weight for row in positive.rows], [Decimal("0.34"), Decimal("0.33"), Decimal("0.33")])
        self.assertEqual([row.adjusted_weight for row in negative.rows], [Decimal("0.66"), Decimal("0.67"), Decimal("0.67")])
        self.assertEqual(positive.adjusted_total, Decimal("1.00"))
        self.assertEqual(negative.adjusted_total, Decimal("2.00"))

    def test_duplicados_se_conservan_por_fila_y_se_exportan_en_orden(self):
        result = self.result([
            "PRECINTO-001;A;100;1", "PRECINTO-001;B;50;1", "PRECINTO-002;C;150;1",
        ], "150,00")
        self.assertEqual(result.source_total, Decimal("300"))
        self.assertEqual([row.source.precinto for row in result.rows], ["PRECINTO-001", "PRECINTO-001", "PRECINTO-002"])
        self.assertEqual([row.adjusted_weight for row in result.rows], [Decimal("50.00"), Decimal("25.00"), Decimal("75.00")])
        self.assertEqual(result.adjusted_total, Decimal("150.00"))
        content = render_ax_csv(result, "OT-0001").decode("cp1252")
        self.assertEqual(content, "OT-0001;PRECINTO-001;50,00\r\nOT-0001;PRECINTO-001;25,00\r\nOT-0001;PRECINTO-002;75,00\r\n")

    def test_validaciones_de_registro(self):
        cases = {
            ";X;1;1": "EMPTY_SEAL",
            "A;X;;1": "EMPTY_WEIGHT",
            "A;X;abc;1": "INVALID_WEIGHT",
            "A;X;-1;1": "NEGATIVE_WEIGHT",
            "A;X;0;1": "ZERO_SOURCE_TOTAL",
            "A;X": "TOO_FEW_FIELDS",
        }
        for message, code in cases.items():
            with self.subTest(code=code):
                source = self.read([message])
                self.assertIn(code, [issue.code for issue in source.issues])
                with self.assertRaises(DomainValidationError):
                    calculate_adjustment(source, "0,00")

    def test_encabezado_filas_vacias_y_estructura_incorrecta(self):
        source = self.read(["A;X;1,00;1", ""])
        self.assertTrue(source.source_format and source.source_format.has_message_header)
        self.assertEqual(len(source.records), 1)
        self.assertEqual(source.ignored_rows[0].reason, "encabezado de mensaje")
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "estructura.xlsx"
            workbook = Workbook()
            workbook.active.append(["A;X;1", "columna B"])
            workbook.save(path)
            workbook.close()
            invalid = read_source_file(path)
        self.assertEqual(invalid.issues[0].code, "INVALID_WORKBOOK_LAYOUT")

    def test_formato_no_soportado_libro_vacio_y_corrupto(self):
        with tempfile.TemporaryDirectory() as directory:
            base = Path(directory)
            unsupported = base / "origen.txt"
            unsupported.write_text("A;X;1", encoding="utf-8")
            empty = base / "vacio.xlsx"
            workbook = Workbook()
            workbook.save(empty)
            workbook.close()
            corrupt = base / "invalido.xlsx"
            corrupt.write_bytes(b"no es un libro Excel")
            self.assertEqual(read_source_file(unsupported).issues[0].code, "UNSUPPORTED_FORMAT")
            self.assertEqual(read_source_file(empty).issues[0].code, "NO_VALID_RECORDS")
            self.assertEqual(read_source_file(corrupt).issues[0].code, "READ_ERROR")

    def test_validacion_peso_final(self):
        total = Decimal("10.00")
        cases = ((None, "EMPTY_FINAL_WEIGHT"), ("abc", "INVALID_FINAL_WEIGHT"), ("-1", "NEGATIVE_FINAL_WEIGHT"), ("1,001", "FINAL_WEIGHT_PRECISION"))
        for value, code in cases:
            with self.subTest(value=value):
                self.assertEqual(validate_final_weight(value, total).issues[0].code, code)
        self.assertTrue(validate_final_weight("11,00", total).is_valid)

    def test_validacion_orden_trabajo_conserva_texto_y_ceros_iniciales(self):
        self.assertEqual(validate_work_order("  000123  ").value, "000123")
        self.assertEqual(validate_work_order(" OT-A/001 ").value, "OT-A/001")
        self.assertEqual(validate_work_order("   ").issues[0].code, "EMPTY_WORK_ORDER")
        self.assertEqual(validate_work_order(None).issues[0].code, "EMPTY_WORK_ORDER")
        result = self.result(["A;X;1,00;1"], "1,00")
        with self.assertRaises(DomainValidationError) as raised:
            render_ax_csv(result, " ")
        self.assertEqual(raised.exception.issues[0].code, "EMPTY_WORK_ORDER")

    def test_csv_ax_tres_columnas_suma_orden_y_bytes(self):
        result = self.result(["B;X;2,00;1", "A;X;1,00;1"], "2,00")
        content = render_ax_csv(result, "OT-001")
        self.assertFalse(content.startswith(b"\xef\xbb\xbf"))
        self.assertTrue(content.endswith(b"\r\n"))
        self.assertEqual(content.decode("cp1252"), "OT-001;B;1,33\r\nOT-001;A;0,67\r\n")
        validate_ax_csv_content(content, result, "OT-001")
        rows = content.decode("cp1252").split("\r\n")[:-1]
        self.assertTrue(all(len(row.split(";")) == 3 for row in rows))

    def test_exportacion_atomica_preserva_el_csv_anterior_si_falla_el_reemplazo(self):
        result = self.result(["A;X;1,00;1"], "1,00")
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "ax.csv"
            output.write_bytes(b"contenido-anterior")
            with patch("suite_pyside6.core.reparto_merma_precintos.os.replace", side_effect=OSError("archivo bloqueado")):
                with self.assertRaises(OSError):
                    write_ax_csv(output, result, "OT-001")
            self.assertEqual(output.read_bytes(), b"contenido-anterior")
            self.assertEqual(list(Path(directory).glob(".ax.csv.*.tmp")), [])

    def test_csv_escapa_la_orden_de_trabajo_con_caracteres_especiales(self):
        result = self.result(["A;X;1,00;1"], "1,00")
        content = render_ax_csv(result, ' OT;"A",001 ').decode("cp1252")
        self.assertEqual(content, '"OT;""A"",001";A;1,00\r\n')

    def test_csv_admite_salto_de_linea_en_orden_de_trabajo(self):
        result = self.result(["A;X;1,00;1"], "1,00")
        content = render_ax_csv(result, "OT-001\nREV-A")
        validate_ax_csv_content(content, result, "OT-001\nREV-A")

    def test_csv_con_cabecera_antepone_orden_de_trabajo(self):
        result = self.result(["A;X;1,00;1"], "1,00")
        export_format = AXCsvFormat(include_header=True)
        content = render_ax_csv(result, "OT-001", export_format).decode("cp1252")
        self.assertEqual(content, "Orden de trabajo;Precinto;Peso ajustado\r\nOT-001;A;1,00\r\n")

    def test_modelo_de_vista_previa_conserva_pesos_y_orden(self):
        result = self.result(["B;X;2,00;1", "A;X;1,00;1"], "2,00")
        preview = build_preview(result)
        self.assertEqual(preview.source_total, Decimal("3.00"))
        self.assertEqual([row.precinto for row in preview.rows], ["B", "A"])
        self.assertEqual(sum((row.peso_ajustado for row in preview.rows), Decimal("0")), Decimal("2.00"))

    def test_csv_injection_se_rechaza_sin_mutar_el_identificador(self):
        result = self.result(["=1+1;X;1,00;1"], "1,00")
        with self.assertRaises(DomainValidationError) as raised:
            render_ax_csv(result, "OT-001")
        self.assertEqual(raised.exception.issues[0].code, "CSV_INJECTION_RISK")

    def test_formato_ax_centralizado(self):
        self.assertEqual(AX_CSV_FORMAT.encoding, "cp1252")
        self.assertEqual(AX_CSV_FORMAT.delimiter, ";")
        self.assertEqual(AX_CSV_FORMAT.decimal_separator, ",")
        self.assertEqual(AX_CSV_FORMAT.precision, 2)
        self.assertEqual(AX_CSV_FORMAT.line_ending, "\r\n")
        self.assertFalse(AX_CSV_FORMAT.include_header)
        self.assertEqual(AX_CSV_FORMAT.headers, ("Orden de trabajo", "Precinto", "Peso ajustado"))

    def test_fac_ignora_vacias_filtra_si_y_reutiliza_formato_ax(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "fac.csv"
            path.write_text(
                "130424;22/7/2026;9:5:29;PREC-1;11.58;7.9;33.07; SI \n"
                ";;;;;;;\n\n\t\n"
                "130424;22/7/2026;9:5:30;PREC-2;12;8.020;33;no\n"
                "130424;22/7/2026;9:5:31;PREC-1;13;8.100;33;SÍ\n",
                encoding="utf-8-sig",
            )
            result = read_fac_files([path])
        self.assertTrue(result.is_valid)
        self.assertEqual(result.ignored_empty_rows, 3)
        self.assertEqual(result.excluded_no_rows, 1)
        self.assertEqual([record.precinto for record in result.records], ["PREC-1", "PREC-1"])
        self.assertEqual(result.total_weight, Decimal("16.000"))
        content = render_ax_csv_records([record.as_ax_record() for record in result.records], "000-OT")
        self.assertEqual(content.decode("cp1252"), "000-OT;PREC-1;7,90\r\n000-OT;PREC-1;8,10\r\n")

    def test_fac_informa_archivo_y_linea_para_si_invalido(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "fac_invalido.csv"
            path.write_text("a;b;c;;1;no-peso;0;SI\n", encoding="utf-8")
            result = read_fac_files([path])
        self.assertFalse(result.is_valid)
        self.assertIn("fac_invalido.csv, línea 1", result.issues[0].message)
        self.assertEqual(result.issues[0].code, "FAC_EMPTY_SEAL")

    def test_fac_conserva_orden_entre_archivos(self):
        with tempfile.TemporaryDirectory() as directory:
            first = Path(directory) / "primero.csv"
            second = Path(directory) / "segundo.csv"
            first.write_text("a;b;c;A;1;1.2;0;SI\na;b;c;B;1;2.3;0;SI\n", encoding="utf-8")
            second.write_text("a;b;c;A;1;3.4;0;SI\n", encoding="utf-8")
            result = read_fac_files([first, second])
        self.assertEqual([record.precinto for record in result.records], ["A", "B", "A"])
        self.assertEqual(result.total_weight, Decimal("6.9"))


if __name__ == "__main__":
    unittest.main()

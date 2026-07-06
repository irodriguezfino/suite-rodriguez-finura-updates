from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
import re
import unicodedata

import openpyxl


EXTENSIONES_EXCEL = {".xlsx", ".xlsm"}


@dataclass(frozen=True)
class RegistroEntrada:
    codigo_articulo: str
    codigo_externo: str
    lote: str
    lote_proveedor: str
    precinto: str
    peso_neto: Decimal
    id_pallet: str


@dataclass(frozen=True)
class RegistroSalida:
    articulo: str
    id_pallet: str
    unidades: int
    kilos: Decimal


@dataclass(frozen=True)
class ExcelDetectado:
    ruta: Path
    tipo: str
    filas: list[RegistroEntrada] | list[RegistroSalida]


@dataclass(frozen=True)
class SalidaGenerada:
    ruta_origen: Path
    nombre_txt: str | None
    lineas: list[str]
    unidades_salida: int
    kilos_salida: Decimal
    jumbos: int


@dataclass(frozen=True)
class ResultadoGeneracion:
    salidas: list[SalidaGenerada]
    lineas: list[str]
    precintos_usados: int
    unidades_salida: int
    kilos_salida: Decimal
    jumbos: int

    def preview_text(self, limit: int = 8) -> str:
        if not self.salidas:
            return "No hay TXT generados."
        blocks: list[str] = []
        for salida in self.salidas:
            sample = "\r\n".join(salida.lineas[:limit])
            if len(salida.lineas) > limit:
                sample += f"\r\n... {len(salida.lineas) - limit} lineas mas"
            blocks.append(f"{texto_nombre_txt(salida.nombre_txt)} ({len(salida.lineas)} lineas)\r\n{sample}")
        return "\r\n\r\n".join(blocks)


@dataclass(frozen=True)
class ExpedicionCarga:
    entrada: ExcelDetectado | None
    salidas: list[ExcelDetectado]
    detectados: list[ExcelDetectado]
    log: list[str]

    def ready(self) -> bool:
        return self.entrada is not None and bool(self.salidas)


def normalizar_texto(valor) -> str:
    texto = unicodedata.normalize("NFKD", str(valor or ""))
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^a-zA-Z0-9]+", " ", texto).strip().lower()
    return re.sub(r"\s+", " ", texto)


def limpiar_codigo(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    texto = str(valor).strip()
    if re.fullmatch(r"\d+(?:\.0+)?", texto):
        return texto.split(".", 1)[0]
    return texto


def limpiar_precinto(valor) -> str:
    return re.sub(r"\D+", "", limpiar_codigo(valor))


def decimal_desde_valor(valor, campo: str) -> Decimal:
    if valor is None or str(valor).strip() == "":
        raise ValueError(f"{campo} vacio")
    if isinstance(valor, str):
        texto = valor.strip()
        if "," in texto:
            texto = texto.replace(".", "").replace(",", ".")
    else:
        texto = str(valor)
    try:
        return Decimal(texto)
    except InvalidOperation as exc:
        raise ValueError(f"{campo} no numerico: {valor}") from exc


def cabeceras_normalizadas(ws) -> tuple[dict[str, int], list[str]]:
    fila = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not fila:
        return {}, []
    originales = [str(v or "").strip() for v in fila]
    return {normalizar_texto(v): i for i, v in enumerate(originales) if str(v or "").strip()}, originales


def indice(cabeceras: dict[str, int], nombre: str) -> int | None:
    return cabeceras.get(normalizar_texto(nombre))


def es_excel_entrada(cabeceras: dict[str, int]) -> bool:
    requeridas = [
        "codigo de articulo",
        "numero de lote",
        "numero de lote del proveedor",
        "peso neto",
        "precinto",
        "id de pallet",
    ]
    return all(indice(cabeceras, nombre) is not None for nombre in requeridas)


def es_excel_salida(cabeceras: dict[str, int]) -> bool:
    requeridas = ["id de pallet", "unidades", "kilos"]
    return all(indice(cabeceras, nombre) is not None for nombre in requeridas)


def leer_excel(ruta: Path) -> ExcelDetectado:
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    try:
        ws = wb.active
        cabeceras, _originales = cabeceras_normalizadas(ws)
        if es_excel_entrada(cabeceras):
            return ExcelDetectado(ruta=ruta, tipo="entrada", filas=leer_filas_entrada(ws, cabeceras))
        if es_excel_salida(cabeceras):
            return ExcelDetectado(ruta=ruta, tipo="salida", filas=leer_filas_salida(ws, cabeceras))
        raise ValueError("no coincide con el formato de entrada ni de salida de expedicion")
    finally:
        wb.close()


def leer_filas_entrada(ws, cabeceras: dict[str, int]) -> list[RegistroEntrada]:
    i_codigo = indice(cabeceras, "codigo de articulo")
    i_externo = indice(cabeceras, "codigo de articulo externo")
    i_lote = indice(cabeceras, "numero de lote")
    i_proveedor = indice(cabeceras, "numero de lote del proveedor")
    i_peso = indice(cabeceras, "peso neto")
    i_precinto = indice(cabeceras, "precinto")
    i_pallet = indice(cabeceras, "id de pallet")
    if None in (i_codigo, i_lote, i_proveedor, i_peso, i_precinto, i_pallet):
        raise ValueError("faltan columnas obligatorias de entrada")

    registros: list[RegistroEntrada] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        precinto = limpiar_precinto(row[i_precinto])
        if not precinto:
            continue
        try:
            peso = decimal_desde_valor(row[i_peso], "Peso neto")
        except ValueError:
            continue
        registros.append(
            RegistroEntrada(
                codigo_articulo=limpiar_codigo(row[i_codigo]),
                codigo_externo=limpiar_codigo(row[i_externo]) if i_externo is not None else "",
                lote=limpiar_codigo(row[i_lote]),
                lote_proveedor=limpiar_codigo(row[i_proveedor]),
                precinto=precinto,
                peso_neto=peso,
                id_pallet=limpiar_codigo(row[i_pallet]),
            )
        )
    if not registros:
        raise ValueError("no hay lineas de entrada con precinto")
    return registros


def leer_filas_salida(ws, cabeceras: dict[str, int]) -> list[RegistroSalida]:
    i_articulo = indice(cabeceras, "codigo de articulo")
    i_pallet = indice(cabeceras, "id de pallet")
    i_unidades = indice(cabeceras, "unidades")
    i_kilos = indice(cabeceras, "kilos")
    if None in (i_pallet, i_unidades, i_kilos):
        raise ValueError("faltan columnas obligatorias de salida")

    registros: list[RegistroSalida] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        unidades_valor = row[i_unidades]
        kilos_valor = row[i_kilos]
        if unidades_valor is None or kilos_valor is None:
            continue
        try:
            unidades = int(decimal_desde_valor(unidades_valor, "Unidades"))
            kilos = decimal_desde_valor(kilos_valor, "Kilos")
        except ValueError:
            continue
        if unidades <= 0:
            continue
        registros.append(
            RegistroSalida(
                articulo=limpiar_codigo(row[i_articulo]) if i_articulo is not None else "",
                id_pallet=limpiar_codigo(row[i_pallet]),
                unidades=unidades,
                kilos=kilos,
            )
        )
    if not registros:
        raise ValueError("no hay jumbos de salida con unidades")
    return registros


def cargar_excels(rutas: list[Path]) -> ExpedicionCarga:
    detectados: list[ExcelDetectado] = []
    log: list[str] = []
    for ruta in rutas:
        if ruta.suffix.lower() not in EXTENSIONES_EXCEL:
            log.append(f"- Ignorado {ruta.name}: no es XLSX/XLSM.")
            continue
        try:
            detectado = leer_excel(ruta)
            detectados.append(detectado)
            log.append(f"- {ruta.name}: detectado como {detectado.tipo} ({len(detectado.filas)} filas utiles).")
        except Exception as exc:
            log.append(f"- ERROR {ruta.name}: {exc}")

    entradas = [d for d in detectados if d.tipo == "entrada"]
    salidas = [d for d in detectados if d.tipo == "salida"]
    entrada = entradas[0] if len(entradas) == 1 else None
    if len(entradas) == 0:
        log.append("- Falta un Excel de entrada.")
    if len(entradas) > 1:
        log.append("- Hay mas de un Excel con formato de entrada. Deja solo uno.")
    if len(salidas) == 0:
        log.append("- Falta al menos un Excel de salida.")
    return ExpedicionCarga(entrada=entrada, salidas=salidas, detectados=detectados, log=log)


def pallets_disponibles(registros: list[RegistroEntrada]) -> list[str]:
    vistos: set[str] = set()
    pallets: list[str] = []
    for registro in registros:
        if registro.id_pallet and registro.id_pallet not in vistos:
            vistos.add(registro.id_pallet)
            pallets.append(registro.id_pallet)
    return pallets


def filtrar_precintos_por_pallets(registros: list[RegistroEntrada], pallets: list[str]) -> list[RegistroEntrada]:
    seleccionados = set(pallets)
    if not seleccionados:
        raise ValueError("Selecciona al menos un Id de pallet de entrada.")
    return [registro for registro in registros if registro.id_pallet in seleccionados and registro.precinto]


def resumen_pivot_entrada(registros: list[RegistroEntrada]) -> list[tuple[str, str, int, Decimal]]:
    resumen: dict[tuple[str, str], tuple[int, Decimal]] = {}
    for registro in registros:
        clave = (registro.codigo_articulo, registro.id_pallet)
        cuenta, peso = resumen.get(clave, (0, Decimal("0")))
        resumen[clave] = (cuenta + 1, peso + registro.peso_neto)
    filas = [(codigo, pallet, cuenta, peso) for (codigo, pallet), (cuenta, peso) in resumen.items()]
    return sorted(filas, key=lambda item: (item[0], item[1]))


def miligramos_desde_kilos(kilos: Decimal) -> int:
    return int((kilos * Decimal(1000)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def repartir_pesos_milesimas(kilos: Decimal, unidades: int) -> list[int]:
    if unidades <= 0:
        raise ValueError("Las unidades deben ser mayores que cero.")
    total_milesimas = miligramos_desde_kilos(kilos)
    base = total_milesimas // unidades
    resto = total_milesimas % unidades
    return [base + 1 if i < resto else base for i in range(unidades)]


def formato_peso_milesimas(milesimas: int) -> str:
    entero = milesimas // 1000
    decimal = abs(milesimas % 1000)
    return f"{entero:02d},{decimal:03d}"


def totales_salida(detectado: ExcelDetectado) -> tuple[int, Decimal, int]:
    salidas = detectado.filas
    unidades = sum(salida.unidades for salida in salidas)  # type: ignore[union-attr]
    kilos = sum((salida.kilos for salida in salidas), Decimal("0"))  # type: ignore[union-attr]
    return unidades, kilos, len(salidas)


def nombre_txt_desde_salida(ruta: Path) -> str | None:
    numeros = re.findall(r"(?<!\d)(\d{6})(?!\d)", ruta.stem)
    if numeros:
        return f"SC{numeros[-1]}.TXT"
    return None


def texto_nombre_txt(nombre: str | None) -> str:
    return nombre if nombre else "Requiere nombre"


def normalizar_nombre_txt_usuario(nombre: str) -> str:
    texto = str(nombre or "").strip().strip('"')
    if not texto:
        raise ValueError("El nombre del TXT no puede estar vacio.")
    if any(caracter in texto for caracter in '<>:"/\\|?*'):
        raise ValueError("El nombre contiene caracteres no permitidos en Windows.")
    if not texto.lower().endswith(".txt"):
        texto += ".TXT"
    return texto


def nombre_unico(destino: Path) -> Path:
    if not destino.exists():
        return destino
    base = destino.stem
    sufijo = destino.suffix
    contador = 2
    while True:
        candidato = destino.with_name(f"{base}_{contador}{sufijo}")
        if not candidato.exists():
            return candidato
        contador += 1


def buscar_combinacion_pallets_exacta(
    pallets: list[str],
    datos: dict[str, tuple[int, Decimal]],
    objetivo: int,
) -> list[str]:
    if objetivo <= 0:
        return []
    candidatos = [(pallet, datos.get(pallet, (0, Decimal("0")))[0]) for pallet in pallets]
    candidatos = [(pallet, cuenta) for pallet, cuenta in candidatos if cuenta > 0]
    exactos = [pallet for pallet, cuenta in candidatos if cuenta == objetivo]
    if exactos:
        return [exactos[0]]

    candidatos.sort(key=lambda item: item[1], reverse=True)
    sufijos = [0] * (len(candidatos) + 1)
    for item_index in range(len(candidatos) - 1, -1, -1):
        sufijos[item_index] = sufijos[item_index + 1] + candidatos[item_index][1]
    mejor: list[str] | None = None

    def buscar(item_index: int, suma: int, seleccion: list[str]) -> bool:
        nonlocal mejor
        if suma == objetivo:
            mejor = list(seleccion)
            return True
        if suma > objetivo or item_index >= len(candidatos):
            return False
        if suma + sufijos[item_index] < objetivo:
            return False
        pallet, cuenta = candidatos[item_index]
        seleccion.append(pallet)
        if buscar(item_index + 1, suma + cuenta, seleccion):
            return True
        seleccion.pop()
        return buscar(item_index + 1, suma, seleccion)

    buscar(0, 0, [])
    return mejor or []


def generar_txts_expedicion(
    entradas_filtradas: list[RegistroEntrada],
    salidas_detectadas: list[ExcelDetectado],
    inicio: datetime | None = None,
) -> ResultadoGeneracion:
    if not salidas_detectadas:
        raise ValueError("Selecciona al menos un Excel de salida.")
    total_unidades = sum(totales_salida(salida)[0] for salida in salidas_detectadas)
    if len(entradas_filtradas) != total_unidades:
        raise ValueError(
            f"Los precintos filtrados ({len(entradas_filtradas)}) no coinciden con las unidades de salida ({total_unidades})."
        )

    inicio = inicio or datetime.now()
    indice_precinto = 0
    segundo = 0
    salidas_generadas: list[SalidaGenerada] = []
    todas_lineas: list[str] = []
    for salida_detectada in salidas_detectadas:
        lineas_salida: list[str] = []
        salidas = salida_detectada.filas
        for salida in salidas:  # type: ignore[assignment]
            pesos = repartir_pesos_milesimas(salida.kilos, salida.unidades)
            if sum(pesos) != miligramos_desde_kilos(salida.kilos):
                raise ValueError(f"No cuadra el peso del jumbo {salida.id_pallet}.")
            for peso in pesos:
                entrada = entradas_filtradas[indice_precinto]
                momento = inicio + timedelta(seconds=segundo)
                lineas_salida.append(
                    ";".join(
                        [
                            entrada.lote_proveedor,
                            momento.strftime("%d/%m/%Y"),
                            momento.strftime("%H:%M:%S"),
                            entrada.codigo_externo,
                            entrada.precinto,
                            entrada.lote,
                            formato_peso_milesimas(peso),
                            "",
                        ]
                    )
                )
                indice_precinto += 1
                segundo += 1
        unidades, kilos, jumbos = totales_salida(salida_detectada)
        salidas_generadas.append(
            SalidaGenerada(
                ruta_origen=salida_detectada.ruta,
                nombre_txt=nombre_txt_desde_salida(salida_detectada.ruta),
                lineas=lineas_salida,
                unidades_salida=unidades,
                kilos_salida=kilos,
                jumbos=jumbos,
            )
        )
        todas_lineas.extend(lineas_salida)

    kilos_total = sum((salida.kilos_salida for salida in salidas_generadas), Decimal("0"))
    jumbos_total = sum(salida.jumbos for salida in salidas_generadas)
    return ResultadoGeneracion(
        salidas=salidas_generadas,
        lineas=todas_lineas,
        precintos_usados=len(entradas_filtradas),
        unidades_salida=total_unidades,
        kilos_salida=kilos_total,
        jumbos=jumbos_total,
    )


def generar_txt_expedicion(
    entradas_filtradas: list[RegistroEntrada],
    salidas: list[RegistroSalida],
    inicio: datetime | None = None,
) -> ResultadoGeneracion:
    salida_virtual = ExcelDetectado(ruta=Path("SC.TXT"), tipo="salida", filas=salidas)
    return generar_txts_expedicion(entradas_filtradas, [salida_virtual], inicio=inicio)


def guardar_txts_expedicion(
    resultado: ResultadoGeneracion,
    destino_dir: Path,
    nombres_manual: dict[str, str] | None = None,
) -> list[Path]:
    destino_dir.mkdir(parents=True, exist_ok=True)
    nombres_manual = nombres_manual or {}
    guardados: list[Path] = []
    for salida in resultado.salidas:
        nombre_txt = salida.nombre_txt or nombres_manual.get(str(salida.ruta_origen))
        if not nombre_txt:
            raise ValueError(f"Falta nombre TXT para {salida.ruta_origen.name}.")
        destino = nombre_unico(destino_dir / normalizar_nombre_txt_usuario(nombre_txt))
        with destino.open("w", encoding="cp1252", newline="") as handle:
            handle.write("\r\n".join(salida.lineas))
            handle.write("\r\n")
        guardados.append(destino)
    return guardados


def texto_decimal(valor: Decimal) -> str:
    texto = format(valor.normalize(), "f")
    return texto.replace(".", ",")

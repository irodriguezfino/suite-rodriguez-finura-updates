"""Dominio para repartir una merma de peso entre precintos y exportar a AX.

El contrato de salida se concentra en :data:`AX_CSV_FORMAT`: orden de trabajo,
precinto y peso ajustado, sin cabecera, ``;``, decimal con coma, cp1252, CRLF
y sin BOM. Los importes de
peso usan :class:`decimal.Decimal` de extremo a extremo.
"""

from __future__ import annotations

import csv
import io
import os
import re
import tempfile
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_FLOOR, ROUND_HALF_UP, localcontext
from pathlib import Path
from typing import Literal
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.workbook import Workbook


Severity = Literal["error"]


@dataclass(frozen=True)
class ValidationIssue:
    code: str
    message: str
    severity: Severity = "error"
    line_number: int | None = None


@dataclass(frozen=True)
class IgnoredRow:
    line_number: int
    reason: str


@dataclass(frozen=True)
class SourceFormat:
    worksheet: str
    column: str
    has_message_header: bool


@dataclass(frozen=True)
class SourceRecord:
    line_number: int
    position: int
    precinto: str
    peso_original: Decimal


@dataclass(frozen=True)
class AXExportRecord:
    """Registro normalizado que entiende el escritor común de AX."""

    precinto: str
    peso_final: Decimal
    source_name: str = ""
    line_number: int | None = None


@dataclass(frozen=True)
class FACRecord:
    articulo: str
    fecha: str
    hora: str
    precinto: str
    peso_inicial: str
    peso_deshuesado: Decimal
    porcentaje_merma: str
    seleccion: str
    archivo_origen: str
    numero_linea: int

    def as_ax_record(self) -> AXExportRecord:
        return AXExportRecord(self.precinto, self.peso_deshuesado, self.archivo_origen, self.numero_linea)


@dataclass(frozen=True)
class FACReadResult:
    records: tuple[FACRecord, ...]
    ignored_empty_rows: int
    excluded_no_rows: int
    issues: tuple[ValidationIssue, ...]

    @property
    def errors(self) -> tuple[ValidationIssue, ...]:
        return self.issues

    @property
    def total_weight(self) -> Decimal:
        return sum((record.peso_deshuesado for record in self.records), Decimal("0"))

    @property
    def is_valid(self) -> bool:
        return bool(self.records) and not self.issues


@dataclass(frozen=True)
class SourceReadResult:
    source_format: SourceFormat | None
    records: tuple[SourceRecord, ...]
    ignored_rows: tuple[IgnoredRow, ...]
    issues: tuple[ValidationIssue, ...]

    @property
    def errors(self) -> tuple[ValidationIssue, ...]:
        return tuple(issue for issue in self.issues if issue.severity == "error")

    @property
    def total_weight(self) -> Decimal:
        return sum((record.peso_original for record in self.records), Decimal("0"))

    @property
    def is_valid(self) -> bool:
        return bool(self.records) and not self.errors

    def require_valid(self) -> None:
        if self.is_valid:
            return
        if self.errors:
            raise DomainValidationError(self.errors)
        raise DomainValidationError((ValidationIssue("NO_VALID_RECORDS", "No hay registros validos."),))


@dataclass(frozen=True)
class AXCsvFormat:
    encoding: str = "cp1252"
    delimiter: str = ";"
    decimal_separator: str = ","
    precision: int = 2
    line_ending: str = "\r\n"
    include_header: bool = False
    bom: bytes = b""
    headers: tuple[str, ...] = ("Orden de trabajo", "Precinto", "Peso ajustado")


AX_CSV_FORMAT = AXCsvFormat()


@dataclass(frozen=True)
class FinalWeightValidation:
    weight: Decimal | None
    issues: tuple[ValidationIssue, ...]

    @property
    def errors(self) -> tuple[ValidationIssue, ...]:
        return tuple(issue for issue in self.issues if issue.severity == "error")

    @property
    def is_valid(self) -> bool:
        return self.weight is not None and not self.errors

    def require_valid(self) -> Decimal:
        if self.is_valid:
            assert self.weight is not None
            return self.weight
        raise DomainValidationError(self.errors)


@dataclass(frozen=True)
class WorkOrderValidation:
    value: str | None
    issues: tuple[ValidationIssue, ...]

    @property
    def errors(self) -> tuple[ValidationIssue, ...]:
        return tuple(issue for issue in self.issues if issue.severity == "error")

    @property
    def is_valid(self) -> bool:
        return self.value is not None and not self.errors

    def require_valid(self) -> str:
        if self.is_valid:
            assert self.value is not None
            return self.value
        raise DomainValidationError(self.errors)


@dataclass(frozen=True)
class AdjustmentRow:
    source: SourceRecord
    exact_weight: Decimal
    adjusted_weight: Decimal
    rounding_units: int


@dataclass(frozen=True)
class AdjustmentResult:
    source_total: Decimal
    final_weight: Decimal
    absolute_loss: Decimal
    loss_percentage: Decimal
    adjustment_factor: Decimal
    rows: tuple[AdjustmentRow, ...]

    @property
    def adjusted_total(self) -> Decimal:
        return sum((row.adjusted_weight for row in self.rows), Decimal("0"))


@dataclass(frozen=True)
class PreviewRow:
    line_number: int
    precinto: str
    peso_original: Decimal
    peso_ajustado: Decimal


@dataclass(frozen=True)
class PreviewModel:
    source_total: Decimal
    final_weight: Decimal
    absolute_loss: Decimal
    loss_percentage: Decimal
    adjustment_factor: Decimal
    rows: tuple[PreviewRow, ...]


class DomainValidationError(ValueError):
    """Error que conserva los problemas mostrables por la futura interfaz."""

    def __init__(self, issues: tuple[ValidationIssue, ...]):
        self.issues = issues
        super().__init__("; ".join(issue.message for issue in issues))


def _is_message_header(value: str) -> bool:
    return value.strip().casefold().startswith("mensaje")


def _parse_decimal(value: str, field_name: str, line_number: int | None = None) -> Decimal:
    text = str(value).strip()
    if not text:
        raise ValueError(f"{field_name} vacio")
    sign = ""
    if text[:1] in "+-":
        sign, text = text[:1], text[1:]
    if not text or not re.fullmatch(r"[0-9.,]+", text):
        raise ValueError(f"{field_name} no numerico")

    comma = "," in text
    dot = "." in text
    if comma and dot:
        if text.rfind(",") > text.rfind("."):
            normalized = text.replace(".", "").replace(",", ".")
        else:
            normalized = text.replace(",", "")
    elif comma:
        normalized = text.replace(",", ".")
    else:
        normalized = text
    try:
        return Decimal(sign + normalized)
    except InvalidOperation as exc:
        suffix = f" en linea {line_number}" if line_number is not None else ""
        raise ValueError(f"{field_name} no numerico{suffix}") from exc


def _message_column_index(rows: list[tuple[object, ...]]) -> int:
    """Localiza la columna que realmente contiene los mensajes de PDA.

    El formato histórico deja el mensaje en A. Las exportaciones actuales de
    precintos ibéricos conservan información y rótulo en A/B y colocan el
    mensaje completo en C. Se elige la columna con más valores delimitados.
    """

    scores: dict[int, int] = {}
    headers: list[int] = []
    for row in rows:
        for index, value in enumerate(row):
            text = "" if value is None else str(value).strip()
            if _is_message_header(text):
                headers.append(index)
            if text.count(";") >= 2:
                scores[index] = scores.get(index, 0) + 1
    if scores:
        return min(scores, key=lambda index: (-scores[index], index))
    if headers:
        return headers[0]
    return 0


def _read_message_workbook(workbook: Workbook) -> SourceReadResult:
    records: list[SourceRecord] = []
    ignored: list[IgnoredRow] = []
    issues: list[ValidationIssue] = []
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    message_column = _message_column_index(rows)
    has_message_header = any(
        _is_message_header("" if value is None else str(value))
        for row in rows
        for value in row
    )

    for line_number, row in enumerate(rows, start=1):
        value = row[message_column] if message_column < len(row) else None
        if value is None or not str(value).strip():
            ignored.append(IgnoredRow(line_number, "fila vacia"))
            continue
        # En el legado, una segunda columna con datos era un Excel mal formado.
        # El formato nuevo usa A/B como metadatos y C para el mensaje, por lo
        # que esa regla solo se mantiene cuando el mensaje está en A.
        if message_column == 0 and any(cell is not None and str(cell).strip() for cell in row[1:]):
            issues.append(
                ValidationIssue(
                    "INVALID_WORKBOOK_LAYOUT",
                    "Cada mensaje debe estar contenido en la primera columna del Excel.",
                    line_number=line_number,
                )
            )
            continue
        message = str(value).strip()
        if _is_message_header(message):
            ignored.append(IgnoredRow(line_number, "encabezado de mensaje"))
            continue
        try:
            fields = next(csv.reader([message], delimiter=";", quotechar='"', strict=True))
        except csv.Error as exc:
            issues.append(ValidationIssue("INVALID_MESSAGE_ROW", f"Mensaje invalido: {exc}", line_number=line_number))
            continue
        if len(fields) < 3:
            issues.append(
                ValidationIssue("TOO_FEW_FIELDS", "El mensaje tiene menos de tres campos.", line_number=line_number)
            )
            continue
        precinto = fields[0].strip()
        if not precinto:
            issues.append(ValidationIssue("EMPTY_SEAL", "El precinto esta vacio.", line_number=line_number))
            continue
        raw_weight = fields[2].strip()
        if not raw_weight:
            issues.append(ValidationIssue("EMPTY_WEIGHT", "El peso esta vacio.", line_number=line_number))
            continue
        try:
            weight = _parse_decimal(raw_weight, "El peso", line_number)
        except ValueError as exc:
            issues.append(ValidationIssue("INVALID_WEIGHT", str(exc), line_number=line_number))
            continue
        if weight < 0:
            issues.append(ValidationIssue("NEGATIVE_WEIGHT", "El peso no puede ser negativo.", line_number=line_number))
            continue
        records.append(SourceRecord(line_number, len(records), precinto, weight))

    total = sum((record.peso_original for record in records), Decimal("0"))
    if records and total == 0:
        issues.append(ValidationIssue("ZERO_SOURCE_TOTAL", "El peso total de origen es cero."))
    if not records and not issues:
        issues.append(ValidationIssue("NO_VALID_RECORDS", "No hay registros validos."))

    source_format = SourceFormat(worksheet.title, get_column_letter(message_column + 1), has_message_header)
    return SourceReadResult(source_format, tuple(records), tuple(ignored), tuple(issues))


def read_source_file(path: Path) -> SourceReadResult:
    """Lee el Excel PDA con precinto primero y peso tercero del mensaje."""

    if path.suffix.lower() != ".xlsx":
        return SourceReadResult(
            None,
            (),
            (),
            (ValidationIssue("UNSUPPORTED_FORMAT", "Solo se admiten libros Excel .xlsx."),),
        )
    try:
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    except (OSError, InvalidFileException, BadZipFile, KeyError, ValueError) as exc:
        return SourceReadResult(None, (), (), (ValidationIssue("READ_ERROR", f"No se pudo leer el fichero: {exc}"),))
    try:
        return _read_message_workbook(workbook)
    finally:
        workbook.close()


def _read_fac_text(path: Path) -> str:
    """Lee exportaciones FAC sin asumir la página de códigos de Windows."""

    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("fac", raw, 0, len(raw), "No se pudo determinar la codificación")


def _is_selected_yes(value: str) -> bool:
    normalized = "".join(
        character for character in unicodedata.normalize("NFD", value.strip().casefold())
        if unicodedata.category(character) != "Mn"
    )
    return normalized == "si"


def _is_selected_no(value: str) -> bool:
    return value.strip().casefold() == "no"


def read_fac_files(paths: list[Path] | tuple[Path, ...]) -> FACReadResult:
    """Lee y valida CSV FAC conservando el orden de archivos y de sus filas.

    Las filas vacías (incluida ``;;;;;;;``) se descartan antes de cualquier
    validación. Solo las filas seleccionadas con SI producen registros AX.
    """

    records: list[FACRecord] = []
    issues: list[ValidationIssue] = []
    ignored_empty_rows = 0
    excluded_no_rows = 0
    for path in paths:
        try:
            text = _read_fac_text(path)
        except (OSError, UnicodeError) as exc:
            issues.append(ValidationIssue("FAC_READ_ERROR", f"{path.name}: no se pudo leer el fichero ({exc})."))
            continue
        for line_number, raw_line in enumerate(text.splitlines(), start=1):
            if not raw_line.strip():
                ignored_empty_rows += 1
                continue
            try:
                fields = next(csv.reader([raw_line], delimiter=";", quotechar='"', strict=True))
            except csv.Error as exc:
                issues.append(ValidationIssue("FAC_INVALID_ROW", f"{path.name}, línea {line_number}: CSV inválido ({exc}).", line_number=line_number))
                continue
            if not any(field.strip() for field in fields):
                ignored_empty_rows += 1
                continue
            if len(fields) != 8:
                issues.append(ValidationIssue("FAC_COLUMN_COUNT", f"{path.name}, línea {line_number}: se esperaban 8 columnas.", line_number=line_number))
                continue
            values = [field.strip() for field in fields]
            if _is_selected_no(values[7]):
                excluded_no_rows += 1
                continue
            if not _is_selected_yes(values[7]):
                # Una marca distinta de SI no participa en la salida; se trata
                # como NO para no convertir un dato operativo normal en alerta.
                excluded_no_rows += 1
                continue
            if not values[3]:
                issues.append(ValidationIssue("FAC_EMPTY_SEAL", f"{path.name}, línea {line_number}: el precinto está vacío.", line_number=line_number))
                continue
            try:
                weight = _parse_decimal(values[5], "El peso deshuesado", line_number)
            except ValueError:
                issues.append(ValidationIssue("FAC_INVALID_WEIGHT", f"{path.name}, línea {line_number}: el peso deshuesado no es numérico.", line_number=line_number))
                continue
            if weight <= 0:
                issues.append(ValidationIssue("FAC_NON_POSITIVE_WEIGHT", f"{path.name}, línea {line_number}: el peso deshuesado debe ser mayor que cero.", line_number=line_number))
                continue
            records.append(FACRecord(*values[:5], weight, *values[6:], path.name, line_number))
    if not records and not issues:
        issues.append(ValidationIssue("FAC_NO_SELECTED_RECORDS", "No hay filas válidas marcadas como SI para exportar."))
    return FACReadResult(tuple(records), ignored_empty_rows, excluded_no_rows, tuple(issues))


def validate_final_weight(
    value: str | Decimal | None,
    source_total: Decimal,
    export_format: AXCsvFormat = AX_CSV_FORMAT,
) -> FinalWeightValidation:
    if value is None or not str(value).strip():
        return FinalWeightValidation(None, (ValidationIssue("EMPTY_FINAL_WEIGHT", "El peso final esta vacio."),))
    try:
        weight = value if isinstance(value, Decimal) else _parse_decimal(str(value), "El peso final")
    except ValueError:
        return FinalWeightValidation(None, (ValidationIssue("INVALID_FINAL_WEIGHT", "El peso final no es numerico."),))
    if weight < 0:
        return FinalWeightValidation(None, (ValidationIssue("NEGATIVE_FINAL_WEIGHT", "El peso final no puede ser negativo."),))
    quantum = Decimal(1).scaleb(-export_format.precision)
    if weight != weight.quantize(quantum):
        return FinalWeightValidation(
            None,
            (
                ValidationIssue(
                    "FINAL_WEIGHT_PRECISION",
                    f"El peso final requiere mas de {export_format.precision} decimales para AX.",
                ),
            ),
        )
    return FinalWeightValidation(weight, ())


def validate_work_order(value: str | None) -> WorkOrderValidation:
    """Valida la orden de trabajo sin alterar su formato significativo."""

    normalized = "" if value is None else str(value).strip()
    if not normalized:
        return WorkOrderValidation(
            None,
            (ValidationIssue("EMPTY_WORK_ORDER", "La orden de trabajo es obligatoria."),),
        )
    return WorkOrderValidation(normalized, ())


def _rounding_order(
    exact_units: list[Decimal],
    rounded_units: list[int],
    records: tuple[SourceRecord, ...],
    positive_residual: bool,
) -> list[int]:
    candidates = range(len(records))
    if not positive_residual:
        candidates = (index for index in candidates if rounded_units[index] > 0)
    return sorted(
        candidates,
        key=lambda index: (
            -(exact_units[index] - exact_units[index].to_integral_value(rounding=ROUND_FLOOR))
            if positive_residual
            else exact_units[index] - exact_units[index].to_integral_value(rounding=ROUND_FLOOR),
            -records[index].peso_original,
            records[index].precinto,
            records[index].position,
        ),
    )


def _reconcile_units(
    exact_units: list[Decimal],
    target_units: int,
    records: tuple[SourceRecord, ...],
) -> list[int]:
    """Conciliacion por mayores restos con desempate estable y comprobable."""

    rounded = [int(units.to_integral_value(rounding=ROUND_HALF_UP)) for units in exact_units]
    residual = target_units - sum(rounded)
    if residual == 0:
        return rounded
    order = _rounding_order(exact_units, rounded, records, positive_residual=residual > 0)
    if not order:
        raise ValueError("No hay filas aptas para conciliar el redondeo.")
    step = 1 if residual > 0 else -1
    for offset in range(abs(residual)):
        index = order[offset % len(order)]
        if rounded[index] + step < 0:
            raise ValueError("La conciliacion produciria un peso negativo.")
        rounded[index] += step
    if sum(rounded) != target_units:
        raise ValueError("La conciliacion no alcanza el peso final.")
    return rounded


def calculate_adjustment(
    source: SourceReadResult,
    final_weight_value: str | Decimal,
    export_format: AXCsvFormat = AX_CSV_FORMAT,
) -> AdjustmentResult:
    """Calcula el reparto proporcional y concilia a la precision de AX."""

    source.require_valid()
    final_validation = validate_final_weight(final_weight_value, source.total_weight, export_format)
    final_weight = final_validation.require_valid()
    quantum = Decimal(1).scaleb(-export_format.precision)
    target_units = int((final_weight / quantum).to_integral_exact())
    with localcontext() as context:
        context.prec = 50
        factor = final_weight / source.total_weight
        exact_weights = [record.peso_original * factor for record in source.records]
        exact_units = [weight / quantum for weight in exact_weights]
        rounded_units = _reconcile_units(exact_units, target_units, source.records)
    rows = tuple(
        AdjustmentRow(
            source=record,
            exact_weight=exact_weight,
            adjusted_weight=(Decimal(units) * quantum),
            rounding_units=units,
        )
        for record, exact_weight, units in zip(source.records, exact_weights, rounded_units, strict=True)
    )
    result = AdjustmentResult(
        source_total=source.total_weight,
        final_weight=final_weight,
        absolute_loss=source.total_weight - final_weight,
        loss_percentage=((source.total_weight - final_weight) / source.total_weight) * Decimal("100"),
        adjustment_factor=factor,
        rows=rows,
    )
    if result.adjusted_total != final_weight:
        raise ValueError("La suma de pesos ajustados no coincide exactamente con el peso final.")
    return result


def build_preview(result: AdjustmentResult) -> PreviewModel:
    return PreviewModel(
        source_total=result.source_total,
        final_weight=result.final_weight,
        absolute_loss=result.absolute_loss,
        loss_percentage=result.loss_percentage,
        adjustment_factor=result.adjustment_factor,
        rows=tuple(
            PreviewRow(
                line_number=row.source.line_number,
                precinto=row.source.precinto,
                peso_original=row.source.peso_original,
                peso_ajustado=row.adjusted_weight,
            )
            for row in result.rows
        ),
    )


def _format_decimal(value: Decimal, export_format: AXCsvFormat) -> str:
    quantum = Decimal(1).scaleb(-export_format.precision)
    text = format(value.quantize(quantum), f".{export_format.precision}f")
    return text.replace(".", export_format.decimal_separator)


def _export_issues(result: AdjustmentResult) -> tuple[ValidationIssue, ...]:
    issues: list[ValidationIssue] = []
    for row in result.rows:
        if row.source.precinto.lstrip().startswith(("=", "+", "-", "@")):
            issues.append(
                ValidationIssue(
                    "CSV_INJECTION_RISK",
                    "El precinto podria interpretarse como formula CSV; no se altera para preservar AX.",
                    line_number=row.source.line_number,
                )
            )
    if result.adjusted_total != result.final_weight:
        issues.append(ValidationIssue("EXPORT_TOTAL_MISMATCH", "El total exportado no coincide con el peso final."))
    return tuple(issues)


def _export_record_issues(records: tuple[AXExportRecord, ...]) -> tuple[ValidationIssue, ...]:
    issues: list[ValidationIssue] = []
    for record in records:
        if record.precinto.lstrip().startswith(("=", "+", "-", "@")):
            location = f"{record.source_name}, línea {record.line_number}: " if record.source_name else ""
            issues.append(ValidationIssue("CSV_INJECTION_RISK", f"{location}el precinto podría interpretarse como fórmula CSV; no se altera para preservar AX.", line_number=record.line_number))
    return tuple(issues)


def render_ax_csv_records(
    records: tuple[AXExportRecord, ...] | list[AXExportRecord],
    work_order: str | None,
    export_format: AXCsvFormat = AX_CSV_FORMAT,
) -> bytes:
    """Escribe registros normalizados con el contrato único de importación AX."""

    normalized_records = tuple(records)
    work_order_validation = validate_work_order(work_order)
    issues = (*work_order_validation.issues, *_export_record_issues(normalized_records))
    if not normalized_records:
        issues = (*issues, ValidationIssue("NO_EXPORT_RECORDS", "No hay registros para exportar."))
    if issues:
        raise DomainValidationError(issues)
    output = io.StringIO(newline="")
    writer = csv.writer(output, delimiter=export_format.delimiter, quotechar='"', quoting=csv.QUOTE_MINIMAL, lineterminator=export_format.line_ending)
    if export_format.include_header:
        writer.writerow(export_format.headers)
    for record in normalized_records:
        writer.writerow((work_order_validation.value, record.precinto, _format_decimal(record.peso_final, export_format)))
    content = export_format.bom + output.getvalue().encode(export_format.encoding, errors="strict")
    validate_ax_csv_records_content(content, normalized_records, work_order_validation.value, export_format)
    return content


def write_ax_csv_records(
    path: Path,
    records: tuple[AXExportRecord, ...] | list[AXExportRecord],
    work_order: str | None,
    export_format: AXCsvFormat = AX_CSV_FORMAT,
) -> None:
    _write_bytes_atomically(path, render_ax_csv_records(records, work_order, export_format))


def render_ax_csv(
    result: AdjustmentResult,
    work_order: str | None,
    export_format: AXCsvFormat = AX_CSV_FORMAT,
) -> bytes:
    """Genera y valida el CSV AX, sin mutar identificadores de precinto."""

    work_order_validation = validate_work_order(work_order)
    issues = (*work_order_validation.issues, *_export_issues(result))
    if issues:
        raise DomainValidationError(issues)
    records = tuple(AXExportRecord(row.source.precinto, row.adjusted_weight, line_number=row.source.line_number) for row in result.rows)
    content = render_ax_csv_records(records, work_order_validation.value, export_format)
    validate_ax_csv_content(content, result, work_order_validation.value, export_format)
    return content


def write_ax_csv(
    path: Path,
    result: AdjustmentResult,
    work_order: str | None,
    export_format: AXCsvFormat = AX_CSV_FORMAT,
) -> None:
    _write_bytes_atomically(path, render_ax_csv(result, work_order, export_format))


def _write_bytes_atomically(path: Path, content: bytes) -> None:
    """Sustituye el CSV solo cuando sus bytes ya se han escrito por completo.

    Así, un error de disco, permisos o reemplazo no corrompe un CSV AX que
    ya existía en la ubicación elegida por la persona usuaria.
    """

    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="wb",
            delete=False,
            dir=path.parent,
            prefix=f".{path.name}.",
            suffix=".tmp",
        ) as temporary_file:
            temporary_path = Path(temporary_file.name)
            temporary_file.write(content)
            temporary_file.flush()
            os.fsync(temporary_file.fileno())
        os.replace(temporary_path, path)
        temporary_path = None
    finally:
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)


def validate_ax_csv_content(
    content: bytes,
    result: AdjustmentResult,
    work_order: str | None,
    export_format: AXCsvFormat = AX_CSV_FORMAT,
) -> None:
    """Relee los bytes exportados y comprueba contrato, filas y suma exacta."""

    if export_format.bom:
        if not content.startswith(export_format.bom):
            raise ValueError("El BOM de exportacion no coincide con el formato.")
        payload = content[len(export_format.bom) :]
    else:
        if content.startswith(b"\xef\xbb\xbf"):
            raise ValueError("El CSV AX no debe incluir BOM UTF-8.")
        payload = content
    text = payload.decode(export_format.encoding, errors="strict")
    if not text.endswith(export_format.line_ending):
        raise ValueError("El CSV AX debe terminar en CRLF.")
    expected_count = len(result.rows) + (1 if export_format.include_header else 0)
    parsed = list(
        csv.reader(
            io.StringIO(text, newline=""),
            delimiter=export_format.delimiter,
            quotechar='"',
            strict=True,
        )
    )
    if len(parsed) != expected_count or any(not row for row in parsed):
        raise ValueError("El CSV AX contiene lineas adicionales o inesperadas.")
    if export_format.include_header:
        if tuple(parsed.pop(0)) != export_format.headers:
            raise ValueError("Las cabeceras del CSV AX no coinciden.")
    normalized_work_order = validate_work_order(work_order).require_valid()
    if any(len(row) != 3 for row in parsed):
        raise ValueError("Cada fila del CSV AX debe tener exactamente tres columnas.")
    if [row[0] for row in parsed] != [normalized_work_order] * len(parsed):
        raise ValueError("La orden de trabajo exportada no coincide con la indicada.")
    expected_seals = [row.source.precinto for row in result.rows]
    exported_seals = [row[1] for row in parsed]
    if exported_seals != expected_seals:
        raise ValueError("Los precintos exportados no coinciden con los registros validos.")
    exported_total = sum((_parse_decimal(row[2], "El peso ajustado") for row in parsed), Decimal("0"))
    if exported_total != result.final_weight:
        raise ValueError("La suma del CSV AX no coincide exactamente con el peso final.")


def validate_ax_csv_records_content(
    content: bytes,
    records: tuple[AXExportRecord, ...] | list[AXExportRecord],
    work_order: str | None,
    export_format: AXCsvFormat = AX_CSV_FORMAT,
) -> None:
    """Verifica los bytes comunes sin depender del cálculo proporcional PDA."""

    normalized_records = tuple(records)
    payload = content
    if export_format.bom:
        if not content.startswith(export_format.bom):
            raise ValueError("El BOM de exportación no coincide con el formato.")
        payload = content[len(export_format.bom) :]
    elif content.startswith(b"\xef\xbb\xbf"):
        raise ValueError("El CSV AX no debe incluir BOM UTF-8.")
    text = payload.decode(export_format.encoding, errors="strict")
    if not text.endswith(export_format.line_ending):
        raise ValueError("El CSV AX debe terminar en CRLF.")
    parsed = list(csv.reader(io.StringIO(text, newline=""), delimiter=export_format.delimiter, quotechar='"', strict=True))
    expected_count = len(normalized_records) + (1 if export_format.include_header else 0)
    if len(parsed) != expected_count or any(not row for row in parsed):
        raise ValueError("El CSV AX contiene líneas adicionales o inesperadas.")
    if export_format.include_header and tuple(parsed.pop(0)) != export_format.headers:
        raise ValueError("Las cabeceras del CSV AX no coinciden.")
    normalized_work_order = validate_work_order(work_order).require_valid()
    if any(len(row) != 3 for row in parsed):
        raise ValueError("Cada fila del CSV AX debe tener exactamente tres columnas.")
    if [row[0] for row in parsed] != [normalized_work_order] * len(parsed):
        raise ValueError("La orden de trabajo exportada no coincide con la indicada.")
    if [row[1] for row in parsed] != [record.precinto for record in normalized_records]:
        raise ValueError("Los precintos exportados no coinciden con los registros válidos.")
    exported_weights = [_parse_decimal(row[2], "El peso ajustado") for row in parsed]
    expected_weights = [record.peso_final.quantize(Decimal(1).scaleb(-export_format.precision)) for record in normalized_records]
    if exported_weights != expected_weights:
        raise ValueError("Los pesos exportados no coinciden con los registros válidos.")
